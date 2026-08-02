import openpyxl
import pandas as pd
import numpy as np
import os

class dataChart:
    def __init__(self):
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(BASE_DIR, "DMCcharts2025.xlsx")
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.ws = self.wb.active

        self.df = pd.read_excel(file_path)

    def createRGBcol(self):
        colors_rgb = []
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=3, max_col=3):
            cell = row[0]
            fill = cell.fill
            if fill and fill.fgColor.type == "rgb":
                argb = fill.fgColor.rgb
                hex_rgb = argb[2:]

                r = int(hex_rgb[0:2], 16)
                g = int(hex_rgb[2:4], 16)
                b = int(hex_rgb[4:6], 16)
                colors_rgb.append((r, g, b))
            else:
                colors_rgb.append(None)

        self.df["RGB"] = colors_rgb
        self._buildIndex()

    def _buildIndex(self):
        """Precompute the chart as arrays so lookups are one vectorised pass
        instead of 589 pandas rows. Built once per process; every lookup below
        reads these and returns the original objects, so results are unchanged."""
        missing = [i for i, c in enumerate(self.df["RGB"]) if c is None]
        if missing:
            raise ValueError(
                f"{len(missing)} DMC row(s) have no RGB fill (first at sheet row "
                f"{missing[0] + 2}). The chart must use solid rgb fills in column C."
            )

        # Kept as the original objects: `Number` mixes ints (350) and strings
        # ("E310"), and callers compare raw values against `usedColor`.
        self._numbers = list(self.df["Number"])
        self._names = list(self.df["Name"])
        self._rgbs = list(self.df["RGB"])
        self._numStr = [str(n) for n in self._numbers]

        self._palette = np.array(self._rgbs, dtype=np.float64)
        self._indexByNumber = {n: i for i, n in enumerate(self._numbers)}
        self._indexByNumStr = {s: i for i, s in enumerate(self._numStr)}

    def get_datas(self):
        return self.df

    def _distances(self, palette, color):
        """Euclidean RGB distance, float64 throughout so the values are
        bit-identical to the original math.sqrt on Python ints."""
        return np.sqrt(((palette - np.asarray(color, dtype=np.float64)) ** 2).sum(axis=1))

    def findClosestColor(self, color, usedColor, colorsList):
        # The original scanned in order keeping strictly-closer matches, which
        # is the global minimum among unused entries, ties going to the first
        # row. np.argmin returns the first minimum, so the two agree.
        if colorsList == []:
            palette = self._palette
            numbers, names, rgbs = self._numbers, self._names, self._rgbs
            dist = self._distances(palette, color)
            if usedColor:
                for n in usedColor:
                    i = self._indexByNumber.get(n)
                    if i is not None:
                        dist[i] = np.inf
        else:
            rgbs = [self.from_Hex_to_Rgb(elem["hex"]) for elem in colorsList]
            numbers = [elem["num"] for elem in colorsList]
            names = [elem["name"] for elem in colorsList]
            dist = self._distances(np.array(rgbs, dtype=np.float64), color)
            if usedColor:
                used = set(usedColor)
                for i, n in enumerate(numbers):
                    if n in used:
                        dist[i] = np.inf

        if dist.size:
            best = int(np.argmin(dist))
            # 254 is the original's starting threshold: nothing farther is ever
            # accepted, and the caller falls back to the error swatch.
            if dist[best] < 254:
                return (numbers[best], names[best], rgbs[best])

        return ("404", "Error", (255, 0, 0))

    def from_Hex_to_Rgb(self, hex):
        hex = hex.lstrip('#')  # supprime le #
        if len(hex) != 6:
            raise ValueError("La couleur hex doit avoir 6 caractères")

        r = int(hex[0:2], 16)
        g = int(hex[2:4], 16)
        b = int(hex[4:6], 16)

        return (r, g, b)

    def findNewColor(self, colors, color):
        # Three closest shades that aren't already on the palette. The original
        # re-sorted a running top-3 on every row; a single stable argsort gives
        # the same three in the same order, including how ties break by row.
        used = {str(c["num"]) for c in colors}
        used.add(str(color["num"]))

        dist = self._distances(self._palette, self.from_Hex_to_Rgb(color["hex"]))
        allowed = np.fromiter((s not in used for s in self._numStr), dtype=bool, count=len(self._numStr))
        candidates = np.flatnonzero(allowed)
        if candidates.size == 0:
            return []

        best = candidates[np.argsort(dist[candidates], kind="stable")[:3]]
        return [(self._numbers[i], self._names[i], self._rgbs[i]) for i in best]

    def addColor(self, colorNum):
        i = self._indexByNumStr.get(str(colorNum))
        if i is None:
            return ("404", "Error: Not Found", (255, 0, 0))
        return (str(self._numbers[i]), self._names[i], self._rgbs[i])
